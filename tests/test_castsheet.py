"""角色候选表 xlsx 解析测试（fixture 用 openpyxl 现场构建）。"""
from __future__ import annotations

import io

import pytest


def png_bytes(color=(200, 30, 30)) -> bytes:
    from PIL import Image as PImage
    buf = io.BytesIO()
    PImage.new("RGB", (64, 64), color).save(buf, "PNG")
    return buf.getvalue()


def build_sheet(path, with_final=True, yellow_cell=None, skip_img_row=None):
    """构建迷你候选表：人物/简介[/定版]/1/2 列 + 2 行角色。返回 path。"""
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "角色候选表"
    headers = ["人物", "简介"] + (["定版"] if with_final else []) + ["1", "2"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    ws.cell(row=2, column=1, value="Evelyn Hart｜22岁｜女主角")
    ws.cell(row=2, column=2, value="简介A")
    ws.cell(row=3, column=1, value="Lucas Vale｜33岁｜男主角")
    ws.cell(row=3, column=2, value="简介B")
    if yellow_cell:  # 如 "C3"
        ws[yellow_cell].fill = PatternFill("solid", start_color="FFFFF258")
    img_col = "C"  # 定版列或无定版时的黄底列都放在 C
    for row in (2, 3):
        if row == skip_img_row:
            continue
        img = XLImage(io.BytesIO(png_bytes((50 * row, 30, 30))))
        ws.add_image(img, f"{img_col}{row}")
    wb.save(path)
    return path


def test_parse_final_column(tmp_path):
    from app import castsheet
    r = castsheet.parse_castsheet(build_sheet(tmp_path / "t.xlsx"))
    assert r["sheet"] == "角色候选表"
    assert len(r["characters"]) == 2
    e, l = r["characters"]
    assert e["name"] == "Evelyn Hart" and e["intro"] == "简介A"
    assert e["image"] and e["ext"] == ".png"
    assert l["name"] == "Lucas Vale"


def test_parse_yellow_fallback(tmp_path):
    from app import castsheet
    # 无定版列：C3 黄底 → 图从黄底格取；第2行无黄底 → 缺图
    r = castsheet.parse_castsheet(build_sheet(tmp_path / "t.xlsx", with_final=False, yellow_cell="C3"))
    assert r["characters"][0]["image"] is None
    assert r["characters"][1]["image"] is not None


def test_parse_missing_image_row(tmp_path):
    from app import castsheet
    r = castsheet.parse_castsheet(build_sheet(tmp_path / "t.xlsx", skip_img_row=3))
    assert r["characters"][1]["image"] is None


def test_parse_no_person_header_raises(tmp_path):
    import openpyxl
    from app import castsheet
    p = tmp_path / "bad.xlsx"
    wb = openpyxl.Workbook()
    wb.active.cell(row=1, column=1, value="随便什么")
    wb.save(p)
    with pytest.raises(castsheet.CastSheetError):
        castsheet.parse_castsheet(p)


def test_name_without_separator(tmp_path):
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    from app import castsheet
    p = tmp_path / "t.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="人物")
    ws.cell(row=1, column=2, value="定版")
    ws.cell(row=2, column=1, value="Lucky")  # 无｜分隔
    ws.add_image(XLImage(io.BytesIO(png_bytes())), "B2")
    wb.save(p)
    r = castsheet.parse_castsheet(p)
    assert r["characters"][0]["name"] == "Lucky"
