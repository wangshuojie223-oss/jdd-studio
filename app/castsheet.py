"""飞书《角色候选表》xlsx 解析：定版列（或黄底格）的图 = 选定角色参考图。"""
from __future__ import annotations

import io
from pathlib import Path

import openpyxl


class CastSheetError(Exception):
    """候选表无法解析。"""


def _is_yellow(rgb: str | None) -> bool:
    """黄色系填充（FFFFFF00 纯黄 / FFFFF258 浅黄等）：R、G 高，B 低且与 R 差距大。"""
    if not rgb or not isinstance(rgb, str) or len(rgb) != 8:
        return False
    try:
        r, g, b = int(rgb[2:4], 16), int(rgb[4:6], 16), int(rgb[6:8], 16)
    except ValueError:
        return False
    return r >= 200 and g >= 180 and b <= 160 and (r - b) >= 60


def _text(v) -> str:
    return str(v).strip() if v is not None else ""


def _find_sheet_and_header(wb):
    """在全部 sheet 前 10 行找「人物」表头。返回 (ws, header_row, {列名: 列号1-based})。"""
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=10):
            cols = {_text(c.value): c.column for c in row if _text(c.value)}
            if "人物" in cols:
                return ws, row[0].row, cols
    raise CastSheetError("找不到「人物」列，请确认导出的是角色候选表（sheet 名："
                         + "/".join(ws.title for ws in wb.worksheets) + "）")


def _image_map(ws) -> dict:
    """{(row, col): 图片字节}（1-based）——按图片锚点归位到单元格。"""
    out = {}
    for img in ws._images:
        try:
            r, c = img.anchor._from.row + 1, img.anchor._from.col + 1
        except AttributeError:
            continue  # AbsoluteAnchor 无单元格归属，跳过
        out.setdefault((r, c), img._data())
    return out


def _detect_ext(data: bytes) -> str:
    from PIL import Image
    fmt = Image.open(io.BytesIO(data)).format
    return ".jpg" if fmt == "JPEG" else ".png"


def parse_castsheet(xlsx_path: Path) -> dict:
    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        raise CastSheetError(f"无法解析该文件（需要飞书导出的 .xlsx）：{e}") from e
    ws, hrow, cols = _find_sheet_and_header(wb)
    final_col = cols.get("定版")
    intro_col = cols.get("简介")
    images = _image_map(ws)

    # 黄底兜底：无定版列时，黄色填充格上的图=选定图
    yellow_cells: set[tuple[int, int]] = set()
    if final_col is None:
        for row in ws.iter_rows(min_row=hrow + 1):
            for c in row:
                if c.fill and c.fill.patternType and _is_yellow(getattr(c.fill.start_color, "rgb", None)):
                    yellow_cells.add((c.row, c.column))
        if not yellow_cells:
            raise CastSheetError("找不到「定版」列，也没有黄底标记的单元格，无法确定选定图")

    characters = []
    for row in ws.iter_rows(min_row=hrow + 1, max_col=max(cols.values())):
        person = _text(row[cols["人物"] - 1].value)
        if not person:
            continue
        name = person.split("｜")[0].strip() or person
        intro = _text(row[intro_col - 1].value) if intro_col else ""
        if final_col is not None:
            data = images.get((row[0].row, final_col))
        else:
            data = next((images[(row[0].row, c)] for (r, c) in yellow_cells
                         if r == row[0].row and (row[0].row, c) in images), None)
        characters.append({
            "name": name, "intro": intro,
            "image": data, "ext": _detect_ext(data) if data else None,
        })
    if not characters:
        raise CastSheetError("「人物」列下没有任何角色行")
    return {"sheet": ws.title, "characters": characters}
